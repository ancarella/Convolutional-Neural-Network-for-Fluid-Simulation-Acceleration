import netgen.gui
from ngsolve import *
import tensorflow as tf
import numpy as np
import openpyxl as xl
import pandas as pd
import pyvista as pv
import pickle
import time
import seaborn as sb
import matplotlib.pyplot as plt
import math
from netgen.geom2d import SplineGeometry
from ngsolve import internal as ngsint


# VISUALIZATION OPTIONS
# ngsint.viewoptions.drawoutline= 1
# ngsint.viewoptions.drawcolorbar= 0
ngsint.visoptions.lineartexture = 1
# ngsint.visoptions.gridsize = 300
# ngsint.visoptions.showsurfacesolution = 1

nu = 0.001  # viscosity

# time-stepping parameters
tau = 0.0001  # step size
tend = 10  # time length of sim
vmax = 0.5
NN_num_of_repeat = 1
NN_insert = 200  # reinserts the NN every ____ time steps
model = tf.keras.models.load_model("FINAL_MODEL.model")

view_NN_output = False
scaledModel = True
saveOutput = 10000
timejump = 0.02 * NN_num_of_repeat
# Newly defined geometry parameters
geoLength = 2
geoHeight = 0.41
numVertx = 256
numVerty = 66
geoVertxSpacing = geoLength / (numVertx - 1)
geoVertySpacing = geoHeight / (numVerty - 1)
side = 0.1
tstop = 5
timerstart = time.time()
output_image = False


geo = SplineGeometry()
geo.AddRectangle((0, 0), (geoLength, geoHeight),
                 bcs=("wall", "outlet", "wall", "inlet"))  # original length,Height are 2,0.41
# geo.AddCircle ( (geoHeight/2, geoHeight/2), r=0.05, leftdomain=0, rightdomain=1, bc="cyl", maxh=0.02)
geo.AddRectangle((math.floor((geoHeight / 2 - side / 2) / geoVertxSpacing) * geoVertxSpacing,
                  math.floor((geoHeight / 2 - side / 2) / geoVertySpacing) * geoVertySpacing), (
                     math.floor((geoHeight / 2 + side / 2) / geoVertxSpacing) * geoVertxSpacing,
                     math.floor((geoHeight / 2 + side / 2) / geoVertySpacing) * geoVertySpacing), leftdomain=0,
                 rightdomain=1, bc="rectangle", maxh=0.02)


mesh = Mesh(geo.GenerateMesh(maxh=0.07))
mesh.Curve(3)

# V = VectorH1(mesh,order=3, dirichlet="wall|cyl|inlet")
V = VectorH1(mesh, order=3, dirichlet="wall|rectangle|inlet")
Q = H1(mesh, order=2)
X = FESpace([V, Q])
u, p = X.TrialFunction()
v, q = X.TestFunction()

stokes = nu * InnerProduct(grad(u), grad(v)) + div(u) * q + div(v) * p - 1e-10 * p * q
a = BilinearForm(X)
a += stokes * dx
a.Assemble()

# nothing here ...
f = LinearForm(X)
f.Assemble()

# gridfunction for the solution
gfu = GridFunction(X)

# parabolic inflow at inlet:
uin = CoefficientFunction((vmax * 4 * y * (0.41 - y) / (0.41 * 0.41), 0))
gfu.components[0].Set(uin, definedon=mesh.Boundaries("inlet"))

# wall boundary conditions
uwall = CoefficientFunction((0, 0))

# solve Stokes problem for initial conditions:
inv_stokes = a.mat.Inverse(X.FreeDofs())

res = f.vec.CreateVector()
res.data = f.vec - a.mat * gfu.vec
gfu.vec.data += inv_stokes * res

# matrix for implicit Euler
mstar = BilinearForm(X)
mstar += SymbolicBFI(u * v + tau * stokes)
mstar.Assemble()
inv = mstar.mat.Inverse(X.FreeDofs(), inverse="sparsecholesky")

# the non-linear term
conv = BilinearForm(X, nonassemble=True)  # convective term
conv += (grad(u) * u) * v * dx
velocity = gfu.components[0]

# for visualization TURN ON LATER
Draw(Norm(gfu.components[0]), mesh, "velocity", sd=3, min=0, max=vmax * 1.25, autoscale=False)
# Draw (gfu.components[1], mesh, "Pressure", sd=3)  # Pressure
# Draw (gfu.components[0][0], mesh, "velocity_x", sd=3)  #X component of velocity
# Draw (gfu.components[0][1], mesh, "velocity_y", sd=3)  #Y component of velocity


# implicit Euler/explicit Euler splitting method:

# Initializes numpy arrays for pressure and velocity

pressure_data = np.ones((numVerty, numVertx))
velocity_data_x = np.ones((numVerty, numVertx))
velocity_data_y = np.ones((numVerty, numVertx))
flow_tensor = np.ones((3, numVerty, numVertx))
# print(flow_tensor.shape)

out_of_bound_counter = 0
for z in range(numVertx * numVerty):
    row = int(z / numVertx)
    column = (z % numVertx)
    if not (mesh.Contains(geoVertxSpacing * column, geoVertySpacing * row)):
        out_of_bound_counter = out_of_bound_counter + 1
xPoints = np.ones(numVerty * numVertx - out_of_bound_counter)
yPoints = np.ones(numVerty * numVertx - out_of_bound_counter)
pointsContained = np.zeros((numVerty, numVertx))
temp_counter = 0
for z in range(numVertx * numVerty):
    row = int(z / numVertx)
    column = (z % numVertx)
    if mesh.Contains(geoVertxSpacing * column, geoVertySpacing * row):
        xPoints[temp_counter] = geoVertxSpacing * column
        yPoints[temp_counter] = geoVertySpacing * row
        pointsContained[row, column] = 1
        temp_counter += 1
    else:
        pointsContained[row, column] = 0

i = 0
t = 0
file_Data_Counter = 0


PNG_count = 0
with TaskManager():
    while t < tend:
        print("t=", t, )

        tic = time.time()
        conv.Apply(gfu.vec, res)
        res.data += a.mat * gfu.vec
        gfu.vec.data -= tau * inv * res

        if i % NN_insert == 0 and i != 0:
            # NOTE ABOUT ELEMENTS (50x50 cells):10201 total, 2601 (51^2) vertices,2500 (50^2) elements
            # ALSO midpoint of lines. 50*51,51*50 sum of all is 10201 = gridfunc
            pressure_data_flat = gfu.components[1](mesh(xPoints, yPoints))
            velocity_data_x_flat = gfu.components[0][0](mesh(xPoints, yPoints))
            velocity_data_y_flat = gfu.components[0][1](mesh(xPoints, yPoints))
            temp_counter = 0
            for z in range(numVertx * numVerty):
                row = int(z / numVertx)
                column = (z % numVertx)
                if pointsContained[row, column] == 1:
                    pressure_data[row, column] = pressure_data_flat[temp_counter]
                    velocity_data_x[row, column] = velocity_data_x_flat[temp_counter]
                    velocity_data_y[row, column] = velocity_data_y_flat[temp_counter]
                    temp_counter += 1
                else:
                    pressure_data[row, column] = 0
                    velocity_data_x[row, column] = 0
                    velocity_data_y[row, column] = 0

            # The code below is used to export CSVs if desired
            # for z in range(numVertx*numVerty): #structmesh.vertices:
            #     #This version does not rely on the L2 interpolation onto the structured mesh. The for loop needs revision though. simply numVertx*numVerty
            #     row = int(z / numVertx)
            #     column = (z % numVertx)
            #     xPoint = geoVertxSpacing*column
            #     yPoint = geoVertySpacing*row
            #     if mesh.Contains(xPoint,yPoint):
            #         #CSV Version
            #         pressure_data[row, column] = gfu.components[1](mesh(xPoint, yPoint))
            #         velocity_data_x[row, column] = gfu.components[0][0](mesh(xPoint, yPoint))
            #         velocity_data_y[row, column] = gfu.components[0][1](mesh(xPoint, yPoint))
            #         #Pickle Version
            #         # flow_tensor[0,row, column] = gfu.components[1](mesh(xPoint, yPoint))
            #         # flow_tensor[1,row, column] = gfu.components[0][0](mesh(xPoint, yPoint))
            #         # flow_tensor[2,row, column] = gfu.components[0][1](mesh(xPoint, yPoint))
            #     else:
            #         #CSV Version
            #         pressure_data[row, column] = 0
            #         velocity_data_x[row, column] = 0
            #         velocity_data_y[row, column] = 0
            #         #Pickle Version
            #         # flow_tensor[0, row, column] = 0
            #         # flow_tensor[1, row, column] = 0
            #         # flow_tensor[2, row, column] = 0

            if scaledModel:
                b = np.sqrt(np.square(velocity_data_x) + np.square(velocity_data_y))
                MaxV = np.amax(b)
                # flow_tensor[0] = pressure_data / np.square(MaxV)
                flow_tensor[0] = pressure_data
                flow_tensor[1] = velocity_data_x / MaxV
                flow_tensor[2] = velocity_data_y / MaxV
            else:
                flow_tensor[0] = pressure_data
                flow_tensor[1] = velocity_data_x
                flow_tensor[2] = velocity_data_y
            flow_tensor = np.swapaxes(flow_tensor, 0, 2)
            flow_tensor = np.swapaxes(flow_tensor, 0, 1)
            flow_tensor = flow_tensor.reshape(1, flow_tensor.shape[0], flow_tensor.shape[1], flow_tensor.shape[2])
            for repeat in range(NN_num_of_repeat):
                if view_NN_output is True:
                    sb.heatmap(MaxV * flow_tensor[0, 0:66, 0:66, 1], square=True,
                               cmap=sb.color_palette("Spectral", as_cmap=True),
                               vmax=vmax * 1.2)
                    plt.show()
                flow_tensor = model.predict([flow_tensor], use_multiprocessing=True)

            if scaledModel:
                flow_tensor[0, :, :, 0] = flow_tensor[0, :, :, 0]
                flow_tensor[0, :, :, 1] = flow_tensor[0, :, :, 1] * MaxV
                flow_tensor[0, :, :, 2] = flow_tensor[0, :, :, 2] * MaxV
            # Visualization
            if view_NN_output is True:
                sb.heatmap(flow_tensor[0, 0:66, 0:66, 1], square=True, cmap=sb.color_palette("Spectral", as_cmap=True),
                           vmax=vmax * 1.2)
                plt.show()
                input("Pause")
            flow_tensor = flow_tensor.astype('float64')
            flow_tensor[0, 0, :, 1] = 0
            flow_tensor[0, 0, :, 2] = 0
            flow_tensor[0, -1, :, 1] = 0
            flow_tensor[0, -1, :, 2] = 0

            vox_flow_tensor_pressure = VoxelCoefficient((0, 0), (geoLength, geoHeight), flow_tensor[0, 0:66, 0:256, 0],
                                                        linear=True)
            vox_flow_tensor_velocity_x = VoxelCoefficient((0, 0), (geoLength, geoHeight),
                                                          flow_tensor[0, 0:66, 0:256, 1], linear=True)
            vox_flow_tensor_velocity_y = VoxelCoefficient((0, 0), (geoLength, geoHeight),
                                                          flow_tensor[0, 0:66, 0:256, 2], linear=True)
            vox_testing = CoefficientFunction((vox_flow_tensor_velocity_x, vox_flow_tensor_velocity_y))
            # Draw(vox_flow_tensor_velocity_x, mesh, "result")
            gfu.components[1].Set(vox_flow_tensor_pressure)
            gfu.components[0].Set((vox_flow_tensor_velocity_x, vox_flow_tensor_velocity_y))

            # Zero the wall velocities
            setdofs = BitArray(V.ndof)
            setdofs.Clear()
            for el in V.Elements(BND):
                if el.mat in ("wall"):
                    for dof in el.dofs:
                        setdofs.Set(dof)
            tmp = gfu.vec.CreateVector()
            tmp.data = gfu.vec
            gfu.components[0].Set(uwall, definedon=mesh.Boundaries("wall"))
            for dnr in range(V.ndof):
                if not (setdofs[dnr]):
                    gfu.vec[dnr] = tmp[dnr]

            # Re-instate inlet boundary conditions
            setdofs = BitArray(V.ndof)
            setdofs.Clear()
            for el in V.Elements(BND):
                if el.mat in ("inlet"):
                    for dof in el.dofs:
                        setdofs.Set(dof)
            tmp = gfu.vec.CreateVector()
            tmp.data = gfu.vec
            gfu.components[0].Set(uin, definedon=mesh.Boundaries("inlet"))
            for dnr in range(V.ndof):
                if not (setdofs[dnr]):
                    gfu.vec[dnr] = tmp[dnr]

            t = t + timejump  # updates the time to account for neural network time jump
            Redraw()
            # input("Inserted soln")
            flow_tensor = np.ones((3, numVerty, numVertx))

        if t >= tstop and output_image is True:
            netgen.gui.Snapshot(w=2000, h=2000, filename="./Square/fastGT" + str(t) + ".png")

            # This code block allows for CSVs to be exported
            # pressure_data_flat = gfu.components[1](mesh(xPoints, yPoints))
            # velocity_data_x_flat = gfu.components[0][0](mesh(xPoints, yPoints))
            # velocity_data_y_flat = gfu.components[0][1](mesh(xPoints, yPoints))
            # #print(pressure_data_flat.shape)
            # temp_counter = 0
            # for z in range(numVertx * numVerty):
            #     row = int(z / numVertx)
            #     column = (z % numVertx)
            #     if (pointsContained[row, column] == 1):
            #         pressure_data[row, column] = pressure_data_flat[temp_counter]
            #         velocity_data_x[row, column] = velocity_data_x_flat[temp_counter]
            #         velocity_data_y[row, column] = velocity_data_y_flat[temp_counter]
            #         temp_counter += 1
            #     else:
            #         pressure_data[row, column] = 0
            #         velocity_data_x[row, column] = 0
            #         velocity_data_y[row, column] = 0
            #
            # wb = xl.Workbook()
            # wb.create_sheet("Data Sheet")
            # # wb.create_sheet("Interpolated Pressure")
            # ws = wb.active
            # filepathP = "t; " + str(t) + "Pressure.csv"
            # filepathVX = "t; " + str(t) + "Velocity X.csv"
            # filepathVY = "t; " + str(t) + "Velocity Y.csv"
            # filepathP2 = "stew interpolated P.csv"
            # dfP = pd.DataFrame(pressure_data)
            # dfVX = pd.DataFrame(velocity_data_x)
            # dfVY = pd.DataFrame(velocity_data_y)
            # # df2 = pd.DataFrame(stewP2)
            # dfP.to_csv("./Square/outputs/" + filepathP, index=False)
            # dfVX.to_csv("./Square/outputs/" + filepathVX, index=False)
            # dfVY.to_csv("./Square/outputs/" + filepathVY, index=False)
            # file_Data_Counter = file_Data_Counter + 1
        i = i + 1
        t = t + tau
        Redraw()

timerstop = time.time()
# netgen.gui.Snapshot(w=2000, h=2000, filename="./Square/fastNN.png")

print("run time : " + str(timerstop - timerstart))
print(i)
print("end of loop")
